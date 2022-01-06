import os
import time
import requests
import configparser
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor

Message = {
    "992107195974": None,
    "992108067660": None,
    "992109150430": None
}

Pool = ProcessPoolExecutor(3)

class Stock_System():
    def get_config_excel(self):
        settings_path = "{}/settings.ini".format(os.path.dirname(os.path.abspath("__file__")))
        config = configparser.ConfigParser()
        config.read(settings_path, encoding='utf-8')
        excel_path = config.get("settings", "excel_path")
        web_hook_url = config.get("settings", "wx_url")
        list1 = [excel_path, web_hook_url]
        return list1


    def read_excel_get_msg(self, excel_path):
        wb = load_workbook(excel_path)
        sheet = wb.worksheets[0]
        list2 = []
        for num in range(2, 5):
            cell = sheet.cell(num, 1)
            res = requests.get(
                url="http://mall.10010.com/goodsdetail/{}.html".format(cell.value)
            )
            text = res.content
            soup = BeautifulSoup(text, features="html.parser")
            area_object = soup.find(name="dl", attrs={"id": "articleAmount"})
            stock = area_object.find(name="span", attrs={"id": "amountChange_id"})
            msg = "商品ID为：{}-{}".format(cell.value, stock.text)
            goods_id = cell.value
            # print(msg)
            now_time = datetime.now()
            string_time = now_time.strftime("%Y-%m-%d %H:%M:%S")
            list2.append([string_time, msg, goods_id])
        return list2


    def send_msg(self, web_hook_url, string_time, msg):
        res = requests.post(
            url=web_hook_url,
            json={
                "msgtype": "text",
                "text": {
                    "content": "报警时间：{}\n报警内容：{}".format(string_time, msg),
                    "mentioned_list": ["@all"]
                }
            }

        )
        res.close()


    def run(self):

        while True:
            excel_path = self.get_config_excel()[0]
            web_hook_url = self.get_config_excel()[1]
            for i in self.read_excel_get_msg(excel_path):
                string_time = i[0]
                msg = i[1]
                ctime = time.time()
                if not Message.get(i[2]):
                    Message[i[2]] = ctime
                    Pool.submit(self.send_msg, web_hook_url, string_time, msg)
                if Message[i[2]] + 20 < ctime:
                    Message[i[2]] = ctime
                    Pool.submit(self.send_msg, web_hook_url, string_time, msg)
                else:
                    continue


if __name__ == "__main__":
    Stock_System().run()